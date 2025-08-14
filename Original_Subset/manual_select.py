#script to convert manual selection to csv

from openpyxl import load_workbook
import csv
import os
import sys

def is_row_highlighted(row):
    """
    Returns True if any cell in the row has a background fill color
    that's not default (white or transparent).
    """
    for cell in row:
        fill = cell.fill
        if fill and fill.fill_type is not None:
            color = fill.start_color.rgb
            if color not in [None, "00000000", "FFFFFFFF"]:
                return True
    return False

def extract_highlighted_rows(input_file, output_file):
    if not os.path.exists(input_file):
        print(f"❌ File not found: {input_file}")
        sys.exit(1)

    try:
        wb = load_workbook(input_file)
        sheetnames = wb.sheetnames
        print("📄 Available sheets:", sheetnames)

        if not sheetnames:
            print("❌ No sheets found in workbook.")
            sys.exit(1)

        ws = wb[sheetnames[0]]
        highlighted_rows = []

        for row in ws.iter_rows():
            if is_row_highlighted(row):
                highlighted_rows.append([cell.value for cell in row])

        if not highlighted_rows:
            print("⚠️ No highlighted rows found.")
        else:
            with open(output_file, mode='w', newline='') as f:
                writer = csv.writer(f)
                writer.writerows(highlighted_rows)

            print(f"✅ Successfully extracted {len(highlighted_rows)} highlighted rows to '{output_file}'.")

    except Exception as e:
        print(f"❌ Error processing file: {e}")
        sys.exit(1)

# ==== EDIT THIS SECTION ====

# Your input Excel file and desired output CSV file
input_excel_file = "utterances_zerosum_GROUND.xlsx"
output_csv_file = "highlighted_rows_zerosum.csv"

# ============================

# Run the function
extract_highlighted_rows(input_excel_file, output_csv_file)
